import pandas as pd
import random
import requests
import itertools
import time
import re

from openpyxl import Workbook
from openpyxl.styles import Font
from HTMS_App.models import *
from django.db.models import Q
from django.contrib.auth.models import User, Group
from datetime import datetime
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.hashers import make_password
from datetime import date, timedelta
from HTMS_App.sms_sender import SendSms, default_message
from .forms import UploadFileForm
from .sqlalchemy_con import SqlAlchemyConnection

from django.http import HttpResponse
from HTMS_App.models import AssetDetails


class Support:
    def __init__(self):
        pass

    def post_to_database(self, request):
        # try:
            date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            technician, status, req_asin_time = self.get_technician_and_status(
                request, date_now
            )
            search_user_selection = request.POST.get("search_user_selection", -1)
            if int(search_user_selection) > 0:
                req_user = Technician.objects.get(
                    user=User.objects.get(id=search_user_selection)
                )
            else:
                self.create_new_user(request)
                req_user = Technician.objects.get(
                    pr_number=request.POST["requester_pr_number"]
                )

            new_serv_req = Requests(
                requester_name=req_user.user.get_full_name(),
                requester_pr_number=req_user.pr_number,
                requester_designation=req_user.designation,
                requester_department=req_user.department,
                requester_email=req_user.user.email,
                requester_extension=req_user.extension_number,
                requester_phone_number=req_user.mobile_number,
                request_type=request.POST["request_type"],
                request_status=status,
                request_mode=request.POST["request_mode"],
                request_priority=request.POST["request_priority"],
                request_category=request.POST["request_category"],
                request_technician=technician,
                subject=request.POST["subject"],
                description=request.POST["description"],
                request_creation_date=date_now,
                request_submitter=request.user,
                last_modified_by=request.user,
                last_modified_date=date_now,
                request_assigned_time=req_asin_time,
                location=request.POST["location"],
            )
            new_serv_req.save()
            if new_serv_req.request_status != "Open":
                self.send_sms_to_technician(technician, new_serv_req)

            # new_serv_req = Requests(
            # requester_name=request.POST["requester_name"],
            # requester_pr_number=request.POST["requester_pr_number"],
            # requester_designation=request.POST["requester_designation"],
            # requester_department=request.POST["requester_department"],
            # requester_email=request.POST["requester_email"],
            # requester_extension=request.POST["requester_extension"],
            # requester_phone_number=request.POST["requester_phone_number"],
        # except Exception as e:
        #     context = {}
        #     context["error"] = [f"❌ Unsuccessful", f"Reason : {e}"]
        #     return context

    def create_new_user(self, request):
        try:
            name = request.POST.get("requester_name", "")
            if not name:
                first_name = request.POST.get("user_first_name", "")
                last_name = request.POST.get("user_last_name", "")
                name = first_name + " " + last_name
            else:
                first_name, *rest = name.split(" ")
                last_name = " ".join(rest)
            employee_data = self.get_employees_details_with_pr_number(
                request.POST.get("requester_pr_number", ""), name
            )
            if employee_data:
                rows = {
                    "department": request.POST.get("requester_department", ""),
                    "user_designation": request.POST.get("requester_designation", ""),
                    "user_extension": request.POST.get("requester_extension", ""),
                }

                self.create_or_get_user_without_request(employee_data, rows)
            if not employee_data:
                raise Exception("No Data was found when searched with SQLAlchemy")
        except:
            name = request.POST.get("requester_name", "")
            if not name:
                first_name = request.POST.get("user_first_name", "")
                last_name = request.POST.get("user_last_name", "")
            else:
                first_name, *rest = name.split(" ")
                last_name = " ".join(rest)

            gender = request.POST.get("user_gender", "")
            pr_num = request.POST.get("requester_pr_number", "")
            email = request.POST.get("requester_email", "")
            department = request.POST.get("requester_department", "")
            designation = request.POST.get("requester_designation", "")
            mobile_number = request.POST.get("requester_phone_number", "")
            extension_number = request.POST.get("requester_extension", "")
            password = make_password(pr_num)
            user, created = User.objects.get_or_create(
                username=pr_num,
                defaults={
                    "password": password,
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": email,
                },
            )

            if created:
                employe_data = Technician(
                    user=User.objects.get(pk=user.id),
                    department=department,
                    designation=designation,
                    pr_number=pr_num,
                    gender=gender,
                    mobile_number=mobile_number,
                    extension_number=extension_number,
                )
                employe_data.save()

                facility_id = request.POST.get("user_facility", "")
                if facility_id:
                    employe_data.facility.set(
                        [FacilityDropdown.objects.get(pk=request.POST["user_facility"])]
                    )
            
                return {
                    "error": "✅ Your Account Has Been Created Successfully!!! ✅. 📢 Your Username & Password is your PR Number. 📢"
                }
            if user:
                if user.first_name == None:
                    user.first_name = first_name
                if user.last_name == None:
                    user.last_name = last_name
                user.save()
                try:
                    tech = Technician.objects.get(user=user)
                except:
                    tech = Technician()
                    tech.user =user
                tech.pr_number = request.POST["requester_pr_number"]
                tech.save()
            else:
                return {
                    "error": f"⚠️ Your Already Have An Account !!! ⚠️. 📢 Enter Your PR Number {pr_num} as Username & Password to Login. 📢"
                }

    # return technician, status, req_asin_time

    def user_update(self, request):
        req_user = Technician.objects.get(pr_number=request.POST["requester_pr_number"])
        req_user.designation = request.POST.get("requester_designation")
        req_user.user.email = request.POST.get("requester_email")
        req_user.extension_number = request.POST.get("requester_extension")
        req_user.mobile_number = request.POST.get("requester_phone_number")
        req_user.save()
        try:
            req_user.user.save()
        except:
            pass

    def send_edit_request_to_db(self, request):
        date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        update_request_inci = Requests.objects.get(pk=request.POST["ticket_pk"])
        technician, status, req_asin_time = self.get_technician_and_status(
            request, date_now, update_request_inci
        )
        update_fields = {
            "requester_name": "requester_name",
            "requester_pr_number": "requester_pr_number",
            "requester_designation": "requester_designation",
            "requester_department": "requester_department",
            "requester_email": "requester_email",
            "requester_extension": "requester_extension",
            "requester_phone_number": "requester_phone_number",
            "request_type": "request_type",
            "request_mode": "request_mode",
            "request_priority": "request_priority",
            "request_category": "request_category",
            "subject": "subject",
            "description": "description",
            "location": "location",
        }
        update_values = {}
        for field, value in update_fields.items():
            try:
                setattr(update_request_inci, field, request.POST[value])
                if request.POST[value] != "" and field != "requester_pr_number":
                    update_values.update(
                        {field.split("_")[-1].title(): request.POST[value]}
                    )

            except KeyError:
                pass

        if update_request_inci.request_status != status:
            update_request_inci.request_status = status
            update_values.update({"Status": status})

        if update_request_inci.request_technician != technician:
            update_request_inci.request_technician = technician
            if (
                technician != None
                and update_request_inci.request_technician.id == technician.id
            ):
                update_values.update(
                    {
                        "Techinican": f"{technician.first_name} {technician.last_name} ({technician.username})"
                    }
                )

        update_request_inci.last_modified_by = request.user
        update_request_inci.last_modified_date = date_now
        update_request_inci.request_assigned_time = req_asin_time
        update_values_in_string = ""
        if update_values != {}:
            for key, value in update_values.items():
                update_values_in_string += (
                    f"\n✔️ {key.capitalize()} Changed To {value} "
                )
            update_request_inci.description += f"\nLast Modified By {request.user.get_full_name()} ({request.user.username}) On {date_now}. Modification: {update_values_in_string}.\n\n"

        update_request_inci.save()
        if update_request_inci.request_status != "Open":
            self.send_sms_to_technician(technician, update_request_inci)
        self.user_update(request)

    def incident_context(self, user_full_name, context):
        context["user_fullname"] = user_full_name
        context["new_inc_req_type"] = NewIncidentRequestType.objects.values()
        context["new_inc_status"] = NewIncidentStatus.objects.values()
        context["new_inc_mode"] = NewIncidentMode.objects.values()
        context["new_inc_priority"] = NewIncidentPriority.objects.values()
        context["new_inc_category"] = (
            NewIncidentCategory.objects.all().order_by("category_name").values()
        )
        technicians_group = Group.objects.get(name="Technicians")
        context["technician"] = technicians_group.user_set.all()
        context["locations_obj"] = Location.objects.values().order_by("location_floor")
        return context

    def display_all_data(request):
        ticket_objects = Requests.objects.all().order_by("-id")
        context = {
            "user_fullname": request.user.get_full_name(),
            "ticket_objects": ticket_objects,
            "header": "All Tickets",
            "link_active_status_all_tickets": "link--active",
        }
        return context

    def non_it_home_content(request):
        user_full_name = request.user.get_full_name()
        technician = Technician.objects.get(user=request.user)
        ticket_objects = Requests.objects.filter(
            requester_pr_number=technician.pr_number
        )
        context = {
            "user_fullname": user_full_name,
            "ticket_objects": ticket_objects,
            "header": f"All Requests By {user_full_name}",
            "link_active_status_all_tickets": "link--active",
            "non_it": True,
        }
        return context
    
    def techs_home_content(request):
        user_full_name = request.user.get_full_name()
        technician = Technician.objects.get(user=request.user)
        ticket_objects = Requests.objects.filter(
            request_technician__technician__pr_number=technician.pr_number
        )
        context = {
            "user_fullname": user_full_name,
            "ticket_objects": ticket_objects,
            "header": f"All Requests Assigned to {user_full_name}",
            "link_active_status_all_tickets": "link--active",
            "techs": True,
        }
        return context

    def search_ticket(request, non_it=False,techs=False):
        search_ticket = ""
        user_full_name = request.user.get_full_name()

        context = {
            "user_fullname": user_full_name,
            "header": "Search Results",
        }
        search_ticket = request.GET.get("search_ticket")
        context["search_ticket"] = search_ticket
        context["page_href"] = f"search_ticket={search_ticket}"
        ticket_objects = (
            Requests.objects.distinct()
            .filter(
                Q(id__icontains=search_ticket)
                | Q(requester_name__icontains=search_ticket)
                | Q(request_status__icontains=search_ticket)
                | Q(request_priority__icontains=search_ticket)
                | Q(request_category__icontains=search_ticket)
                | Q(request_technician__username__icontains=search_ticket)
                | Q(subject__icontains=search_ticket)
                | Q(request_creation_date__icontains=search_ticket)
                | Q(location=search_ticket)
            )
            .order_by("-id")
        )

        if non_it:
            technician = Technician.objects.get(user=request.user)
            ticket_objects = ticket_objects.filter(
                Q(requester_pr_number=technician.pr_number)
            )

        

        if techs:
            technician = Technician.objects.get(user=request.user)
            ticket_objects = ticket_objects.filter(
                Q(request_technician__technician__pr_number=technician.pr_number)
            )

        if not ticket_objects:
            context["error"] = ["No data found!!!", "Please refine your search."]

        else:
            context["ticket_objects"] = ticket_objects

        return context, ticket_objects

    def tickets_to_handle(request, non_it=False,techs=False):
        tickets_to_handle = ""
        user_full_name = request.user.get_full_name()

        context = {
            "user_fullname": user_full_name,
            "header": "Tickets to handle",
            "link_active_status_tickets_to_handle": "link--active",
        }
        tickets_to_handle = request.GET.get("tickets_to_handle")
        context["page_href"] = f"tickets_to_handle={tickets_to_handle}"

        if non_it:
            context["error"] = [
                f" {user_full_name} is not Authorised to View This Page. "
            ]
            return context, None

        ticket_objects = (
            Requests.objects.distinct()
            .filter(Q(request_status__icontains=tickets_to_handle))
            .order_by("-id")
        )

        if not ticket_objects:
            context["error"] = ["No More Tickets To Handle."]

        else:
            context["ticket_objects"] = ticket_objects

        return context, ticket_objects

    def my_open_ticket(request, non_it=False,techs=False):
        my_open_ticket = request.GET.get("my_open_ticket", "open")
        user_pk = request.user.id
        context = {
            "user_fullname": request.user.get_full_name(),
            "header": "My Open Tickets",
            "link_active_status_my_open_ticket": "link--active",
            "page_href": f"my_open_ticket={my_open_ticket}",
        }

        if non_it:
            technician = Technician.objects.get(user=request.user)
            ticket_objects = (
                Requests.objects.distinct()
                .filter(
                    request_status__icontains=my_open_ticket,
                    requester_pr_number=technician.pr_number,
                )
                .order_by("-id")
            )
        else:
            ticket_objects = (
                Requests.objects.distinct()
                .filter(
                    request_status__icontains=my_open_ticket,
                    request_submitter__id=user_pk,
                )
                .order_by("-id")
            )
        if not ticket_objects:
            context["error"] = ["You don't have any Open Tickets."]
        else:
            context["ticket_objects"] = ticket_objects

        return context, ticket_objects

    def my_tick_seven_days(request, non_it=False,techs=False):
        my_tick_svn_days = request.GET.get("my_tick_svn_days")
        context = {
            "user_fullname": request.user.get_full_name(),
            "header": "My tickets in last 7 days",
            "link_active_status_my_tick_seven_days": "link--active",
            "page_href": f"my_tick_svn_days={my_tick_svn_days}",
        }

        today = datetime.now()
        svn_day = today - timedelta(days=7)
        if non_it:
            technician = Technician.objects.get(user=request.user)
            ticket_objects = (
                Requests.objects.distinct()
                .filter(
                    requester_pr_number=technician.pr_number,
                    request_creation_date__range=[svn_day, today],
                )
                .order_by("-id")
            )
        else:
            ticket_objects = (
                Requests.objects.distinct()
                .filter(
                    request_submitter__id=request.user.id,
                    request_creation_date__range=[svn_day, today],
                )
                .order_by("-id")
            )

        if not ticket_objects:
            context["error"] = ["You don't have any Tickets In Last Seven Days."]

        else:
            context["ticket_objects"] = ticket_objects

        return context, ticket_objects

    def tickets_open(request):
        tickets_open = ""
        user_full_name = request.user.get_full_name()

        context = {
            "user_fullname": user_full_name,
            "header": "Open Tickets",
            "link_active_status_tickets_open": "link--active",
        }
        tickets_open = request.GET.get("open")
        context["page_href"] = f"open={tickets_open}"
        ticket_objects = (
            Requests.objects.distinct()
            .filter(Q(request_status__icontains="open"))
            .order_by("-id")
        )

        if not ticket_objects:
            context["error"] = ["No Open Tickets Left."]

        else:
            context["ticket_objects"] = ticket_objects

        return context, ticket_objects

    def tickets_assigned(self, request, non_it=False,techs=False):
        
        tickets_assigned = request.GET.get("assigned")

        context = {
            "user_fullname": request.user.get_full_name(),
            "header": "Assigned Tickets",
            "link_active_status_assigned": "link--active",
            "page_href": f"assigned={tickets_assigned}",
        }

        ticket_objects = (
            Requests.objects.distinct()
            .filter(Q(request_status__icontains="assigned"))
            .order_by("-id")
        )
        if non_it:
            ticket_objects = self.non_it_filter(ticket_objects, request)
        
        if techs:
            ticket_objects = self.techs_filter(ticket_objects, request)

        if not ticket_objects:
            context["error"] = ["No Assigned Tickets to Anyone."]

        else:
            context["ticket_objects"] = ticket_objects

        return context, ticket_objects

    def tickets_closed(self, request, non_it=False,techs=False):
        tickets_closed = request.GET.get("closed")
        user_full_name = request.user.get_full_name()

        context = {
            "user_fullname": user_full_name,
            "header": "Closed Tickets",
            "link_active_status_closed": "link--active",
            "page_href": f"closed={tickets_closed}",
        }
        ticket_objects = (
            Requests.objects.distinct()
            .filter(Q(request_status__icontains="Closed"))
            .order_by("-id")
        )
        if non_it:
            ticket_objects = self.non_it_filter(ticket_objects, request)
            
        if techs:
            ticket_objects = self.techs_filter(ticket_objects, request)

        if not ticket_objects:
            context["error"] = ["No Closed Tickets Found."]

        else:
            context["ticket_objects"] = ticket_objects

        return context, ticket_objects

    def tickets_on_hold(self, request, non_it=False,techs=False):
        tickets_on_hold = request.GET.get("on_hold")
        context = {
            "user_fullname": request.user.get_full_name(),
            "header": "On Hold Tickets",
            "link_active_status_on_hold": "link--active",
            "page_href": f"on_hold={tickets_on_hold}",
        }

        ticket_objects = (
            Requests.objects.distinct()
            .filter(Q(request_status__icontains="On Hold"))
            .order_by("-id")
        )
        if non_it:
            ticket_objects = self.non_it_filter(ticket_objects, request)
        if techs:
            ticket_objects = self.techs_filter(ticket_objects, request)
        if not ticket_objects:
            context["error"] = ["No Tickets On Hold."]

        else:
            context["ticket_objects"] = ticket_objects

        return context, ticket_objects

    def wait_for_feedback(self, request, non_it=False,techs=False):
        wait_for_fback = request.GET.get("wait_for_fback")
        context = {
            "user_fullname": request.user.get_full_name(),
            "header": "Waiting For Feedback",
            "link_active_status_wait_for_fback": "link--active",
            "page_href": f"wait_for_fback={wait_for_fback}",
        }

        ticket_objects = (
            Requests.objects.distinct()
            .filter(Q(request_status__icontains="Waiting for user feedback"))
            .order_by("-id")
        )

        if non_it:
            ticket_objects = self.non_it_filter(ticket_objects, request)
        if techs:
            ticket_objects = self.techs_filter(ticket_objects, request)
        if not ticket_objects:
            context["error"] = ["No Waiting For Feedback Tickets Found!!!."]

        else:
            context["ticket_objects"] = ticket_objects

        return context, ticket_objects

    def tickets_unresolved(self, request, non_it=False,techs=False):
        tick_unresolved = request.GET.get("unresolved")

        context = {
            "user_fullname": request.user.get_full_name(),
            "header": "Unresolved Tickets",
            "link_active_status_unresolved": "link--active",
            "page_href": f"Unresolved={tick_unresolved}",
        }

        ticket_objects = (
            Requests.objects.distinct()
            .filter(Q(request_status__icontains="Unresolved"))
            .order_by("-id")
        )

        if non_it:
            ticket_objects = self.non_it_filter(ticket_objects, request)
        if techs:
            ticket_objects = self.techs_filter(ticket_objects, request)
        if not ticket_objects:
            context["error"] = ["No Unresolved Tickets Found!!!."]

        else:
            context["ticket_objects"] = ticket_objects

        return context, ticket_objects

    def paginator(request, context, paginate_name="ticket_objects"):
        page = request.GET.get("page")
        # results = 14
        results = 16
        paginator = Paginator(context[paginate_name], results)
        # context["paginator"] = paginator

        if not page:
            page = 1
        left_index = int(page) - 4
        if left_index < 1:
            left_index = 1

        right_index = int(page) + 5
        if right_index > paginator.num_pages:
            right_index = paginator.num_pages + 1

        custom_page_range = range(left_index, right_index)
        context["custom_page_range"] = custom_page_range
        try:
            context[paginate_name] = paginator.page(page)
        except PageNotAnInteger:
            page = 1
            context[paginate_name] = paginator.page(page)
        except EmptyPage:
            page = paginator.num_pages
            context[paginate_name] = paginator.page(page)

        return context

    def humanize_datetime(self, time_in_seconds):
        time = timedelta(seconds=time_in_seconds)
        days = time.days
        hours, remainder = divmod(time.seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        time_in_between = {
            "days": days,
            "hours": hours,
            "minutes": minutes,
            "seconds": seconds,
        }
        return time_in_between

    def calculate_time_in_between(self, recent_time, earlier_time):
        try:
            recent_time = recent_time.replace(microsecond=0)
        except:
            return None
        try:
            earlier_time = earlier_time.replace(microsecond=0)
        except:
            return None
        if recent_time == earlier_time:
            return {"seconds": 0}
        try:
            between_time = recent_time - earlier_time
        except:
            return None
        if between_time:
            between_time = self.humanize_datetime(between_time.total_seconds())
        return between_time

    def update_ticket(self, request, pk):
        ticket_edit_objects = Requests.objects.get(pk=pk)
        context = {
            "ticket_edit_objects": ticket_edit_objects,
            "incident_header": f"Update Ticket #{pk}",
        }
        context["submit_type"] = context["incident_header"]
        context = self.incident_context(request.user.get_full_name(), context)
        if ticket_edit_objects.request_assigned_time:
            time_between_assign = self.calculate_time_in_between(
                ticket_edit_objects.request_assigned_time,
                ticket_edit_objects.request_creation_date,
            )
            context["time_between_assign"] = time_between_assign
        if ticket_edit_objects.request_resolved_time:
            time_between_resolve = self.calculate_time_in_between(
                ticket_edit_objects.request_resolved_time,
                ticket_edit_objects.request_assigned_time,
            )
            context["time_between_resolve"] = time_between_resolve
        if ticket_edit_objects.request_closed_time:
            time_between_close = self.calculate_time_in_between(
                ticket_edit_objects.request_closed_time,
                ticket_edit_objects.request_assigned_time,
            )
            context["time_between_close"] = time_between_close
        if ticket_edit_objects.request_closed_time:
            total_tat_time = self.calculate_time_in_between(
                ticket_edit_objects.request_closed_time,
                ticket_edit_objects.request_creation_date,
            )
            context["total_tat_time"] = total_tat_time

        return context

    def get_technician_and_status(self, request, date_now, ticket_edit_objects=None):
        technician_id = request.POST.get("request_technician")
        technician = User.objects.filter(pk=technician_id).first()
        status = "Open"
        req_asin_time = None
        if technician:
            status = "Assigned"
            if ticket_edit_objects and ticket_edit_objects.request_technician:
                status = ticket_edit_objects.request_status
            req_asin_time = date_now

        return technician, status, req_asin_time


    def non_it_filter(self, ticket_objects, request):
        technician = Technician.objects.get(user=request.user)
        ticket_objects = ticket_objects.filter(requester_pr_number=technician.pr_number)
        return ticket_objects
    
    def techs_filter(self, ticket_objects, request):
        technician = Technician.objects.get(user=request.user)
        ticket_objects = ticket_objects.filter(request_technician__technician__pr_number=technician.pr_number)
        return ticket_objects

    def get_assets_creation_context(self, request):
        context = {
            "user_fullname": request.user.get_full_name(),
            "link_active_status_create_new_asset": "link--active",
            "asset_header": "Create A New Asset",
            "create_assest": True,
        }
        return context

    def create_new_asset_type(self, request):
        try:
            image_url_api = self.get_categories(request.POST.get("asset_name"), 1)
            image_url_api = image_url_api["icons"][0]["raster_sizes"][-1]["formats"][
                -1
            ]["preview_url"]
        except:
            num = random.randint(1, 15)

            image_url_api = f"http://127.0.0.1:8000/static/HTMS_App/asset{num}.png"

        date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_asset = Assets(
            asset_name=request.POST.get("asset_name").title(),
            asset_type=request.POST.get("asset_type"),
            asset_description=request.POST.get("description"),
            asset_creation_date=date_now,
            asset_creator=User.objects.get(id=request.user.id),
            image_url=image_url_api,
        )
        try:
            new_asset.save()
        except Exception as e:
            pass

    def get_categories(self, search_key, number_of_res):
        BASE_ENDPOINT = "https://api.iconfinder.com/v4/"
        API_SECRET = "NlhtuuCDXDjWp02MEBMqmWLzXus7jYVxicyPoISuo7aT3BvZJaKXyVO7ecvRuEmu"  # Keep this secret
        url = f"https://api.iconfinder.com/v4/icons/search?query={search_key}&count={number_of_res}"
        # Create the categories endpoint
        categories_endpoint = BASE_ENDPOINT + "categories"

        # Create the authorization header (and any additional ones if needed)
        headers = {"Authorization": "Bearer " + API_SECRET}

        # Make the GET request.
        response = requests.get(url, headers=headers)

        return response.json()

    def get_asset_head_objects(self, context):
        all_assets_heads = (
            Assets.objects.all().order_by("asset_name").values().order_by("asset_name")
        )
        context["all_assets_heads"] = all_assets_heads
        return context

    def get_inventory_home_context(self, request):
        # random_number = random.randint(1, 15)
        asset_objects = AssetDetails.objects.all().order_by("-id")
        context = {
            "user_fullname": request.user.get_full_name(),
            "header": "Assets Summary",
            "link_active_status_all_assests": "link--active",
            # "random_number": random_number,
            "asset_objects": asset_objects,
        }
        context = self.get_asset_head_objects(context)

        return context

    def add_quantity_to_asset(self, request):
        facilities = FacilityDropdown.objects.all().order_by("facility_name").values()
        asset_status = AssetStatus.objects.all().order_by("status_name").values()
        context = {
            "asset_header": "Add Quantity To An Asset",
            "add_quantity": True,
            "facilities": facilities,
            "asset_status": asset_status,
        }
        context = self.incident_context(request.user.get_full_name(), context)
        context = self.get_asset_head_objects(context)
        return context

    def user_creation_or_updation(self, request):
        search_user_selection = request.POST.get("search_user_selection", -1)
        pr_num = request.POST.get("requester_pr_number")

        if int(search_user_selection) > 0:
            user = User.objects.get(pk=search_user_selection)
            return user

        if int(search_user_selection) == 0 or pr_num:
            name = request.POST.get("requester_name", "")
            first_name, *rest = name.split(" ")
            last_name = " ".join(rest)
            password = make_password(pr_num)

            try:
                employee_data = self.get_employees_details_with_pr_number(pr_num, name)
                if employee_data:
                    rows = {
                        "department": request.POST.get("requester_department", ""),
                        "user_designation": request.POST.get(
                            "requester_designation", ""
                        ),
                        "user_extension": request.POST.get("requester_extension", ""),
                    }

                    user = self.create_or_get_user_without_request(employee_data, rows)
                if not employee_data:
                    raise Exception("No Data was found when searched with SQLAlchemy")
                else:
                    return user
            except:
                user, created = User.objects.get_or_create(
                    username=pr_num,
                    defaults={
                        "password": password,
                        "first_name": first_name,
                        "last_name": last_name,
                        "email": request.POST.get("requester_email"),
                    },
                )

                if created:
                    user = User.objects.get(pk=user.id)
                    employe_data = Technician(
                        user=user,
                        department=request.POST.get("requester_department"),
                        designation=request.POST.get("requester_designation"),
                        pr_number=pr_num,
                        mobile_number=request.POST.get("requester_phone_number")[:10],
                        extension_number=request.POST.get("requester_extension")[:10],
                    )
                    employe_data.save()

                return user

    def search_ticket_or_create(self, request, user):
        search_ticket_selection = request.POST.get("search_ticket_selection", -1)
        req_type = request.POST.get("request_type")
        if int(search_ticket_selection) > 0:
            search_ticket_obj = Requests.objects.get(pk=search_ticket_selection)
            return search_ticket_obj

        if int(search_ticket_selection) == 0 or req_type:
            date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            technician, status, req_asin_time = self.get_technician_and_status(
                request, date_now
            )
            new_serv_req = Requests(
                requester_name=f"{user.first_name} {user.last_name}",
                requester_pr_number=user.technician.pr_number,
                requester_designation=user.technician.designation,
                requester_department=user.technician.department,
                requester_email=user.email,
                requester_extension=user.technician.extension_number,
                requester_phone_number=user.technician.mobile_number,
                request_type=req_type,
                request_status=status,
                request_mode=request.POST["request_mode"],
                request_priority=request.POST["request_priority"],
                request_category=request.POST["request_category"],
                request_technician=technician,
                subject=request.POST["subject"],
                request_creation_date=date_now,
                request_submitter=request.user,
                last_modified_by=request.user,
                last_modified_date=date_now,
                request_assigned_time=req_asin_time,
                location=request.POST["location"],
            )
            new_serv_req.save()
            return new_serv_req

    def post_assest_quantity_addition(self, request):
        date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        asset_detail = AssetDetails(
            asset_name=Assets.objects.get(pk=request.POST.get("add_asset_name")),
            brand=request.POST.get("asset_brand"),
            model_name=request.POST.get("model_name"),
            model_number=request.POST.get("model_number"),
            serial_number=request.POST.get("serial_number"),
            date_of_purchase=datetime.strptime(
                request.POST.get("date_of_purchase"), r"%Y-%m-%d"
            ),
            date_added=date_now,
            expiration_date=request.POST.get("expiration_date")
            if request.POST.get("expiration_date") != ""
            else None,
            current_status=request.POST.get("asset_current_status"),
            description=request.POST.get("description"),
            facility=FacilityDropdown.objects.get(
                pk=request.POST.get("asset_facility")
            ),
            added_by=request.user,
        )
        user = self.user_creation_or_updation(request)
        if user:
            asset_detail.asset_user = user

            ticket_req_obj = self.search_ticket_or_create(request, user)
            asset_detail.assign_to_ticket = Requests.objects.get(pk=ticket_req_obj.id)
        asset_detail.save()

    def update_asset(self, request, pk):
        context = self.add_quantity_to_asset(request)
        asset_details_objects = AssetDetails.objects.get(pk=pk)
        context["asset_details_objects"] = asset_details_objects
        context[
            "asset_header"
        ] = f"Update Asset #{pk} {asset_details_objects.brand} - {asset_details_objects.asset_name.asset_name}"
        return context

    def send_edit_request_to_db_asset(self, request):
        date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        user_instance = User.objects.filter(pk=request.user.id).first()
        facility_instance = FacilityDropdown.objects.filter(
            id=request.POST.get("asset_facility")
        ).first()

        update_asset_details = AssetDetails.objects.get(
            pk=request.POST["asset_detail_pk"]
        )
        update_fields = {
            "brand": "asset_brand",
            "model_name": "model_name",
            "model_number": "model_number",
            "serial_number": "serial_number",
            "date_of_purchase": "date_of_purchase",
            "current_status": "asset_current_status",
            "facility": facility_instance,
            "description": "description",
            "expiration_date": "expiration_date",
        }
        update_values = {}

        for field, value in update_fields.items():
            try:
                if field == "facility":
                    if update_asset_details.facility != value:
                        setattr(update_asset_details, field, value)
                        update_values.update({field.replace("_", " ").title(): value})

                if field == "description":
                    setattr(update_asset_details, field, request.POST[value])

                if field == "current_status":
                    if update_asset_details.current_status != request.POST[value]:
                        setattr(update_asset_details, field, request.POST[value])
                        update_values.update(
                            {field.replace("_", " ").title(): request.POST[value]}
                        )
                else:
                    if request.POST[value] == "":
                        pass
                    else:
                        setattr(update_asset_details, field, request.POST[value])
                        update_values.update(
                            {field.replace("_", " ").title(): request.POST[value]}
                        )

            except KeyError:
                pass

        update_asset_details.last_modified_by = user_instance
        update_asset_details.last_modified_date = date_now
        user = self.user_creation_or_updation(request)
        if user:
            if update_asset_details.asset_user != user:
                update_asset_details.asset_user = user
                update_values.update(
                    {"User Changed To ": f"{user.get_full_name()} ({user.username})"}
                )

            ticket_req_obj = self.search_ticket_or_create(request, user)
            if ticket_req_obj:
                update_asset_details.assign_to_ticket = Requests.objects.get(
                    pk=ticket_req_obj.id
                )
                {"Ticket Assigned To ": f"{user.get_full_name()} ({user.username})"}
        update_values_in_string = ""

        if update_values != {}:
            for key, value in update_values.items():
                update_values_in_string += (
                    f"\n✔️ {key.capitalize()} Changed To {value} "
                )
            update_asset_details.description += f"\nLast Modified By {request.user.get_full_name()} ({request.user.username}) On {date_now}. Modification: {update_values_in_string}.\n\n"

        update_asset_details.save()

    def search_assets(self, request):
        search_asset = ""
        context = {
            "user_fullname": request.user.get_full_name(),
            "header": "Search Results",
        }
        search_asset = request.GET.get("search_asset")
        context["search_asset"] = search_asset
        context["page_href"] = f"search_asset={search_asset}"
        context = self.get_asset_head_objects(context)
        if search_asset.isdigit():
            search_with_id = int(search_asset)

        else:
            search_with_id = 0
        asset_objects = (
            AssetDetails.objects.distinct()
            .filter(
                Q(asset_name__asset_name__icontains=search_asset)
                | Q(brand__icontains=search_asset)
                | Q(model_name__icontains=search_asset)
                | Q(model_number__icontains=search_asset)
                | Q(serial_number__icontains=search_asset)
                | Q(date_of_purchase__icontains=search_asset)
                | Q(date_added__icontains=search_asset)
                | Q(current_status__icontains=search_asset)
                | Q(description__icontains=search_asset)
                | Q(facility__facility_name__icontains=search_asset)
                | Q(asset_user__username__icontains=search_asset)
                | Q(asset_user__first_name__icontains=search_asset)
                | Q(asset_user__last_name__icontains=search_asset)
                | Q(asset_user__technician__pr_number__icontains=search_asset)
                | Q(assign_to_ticket__id=search_with_id)
            )
            .order_by("-id")
        )

        if not asset_objects:
            context["error"] = ["No data found!!!", "Please refine your search."]

        else:
            context["asset_objects"] = asset_objects

        return context, asset_objects

    def filter_by_asset_id(self, request):
        context = {
            "user_fullname": request.user.get_full_name(),
            "link_active_status": "link--active",
        }

        asset_id = request.GET.get("asset_id")
        context["page_href"] = f"asset_id={asset_id}"
        asset_objects = (
            AssetDetails.objects.distinct()
            .filter(asset_name__id=asset_id)
            .order_by("-id")
        )
        context = self.get_asset_head_objects(context)
        if not asset_objects:
            context["error"] = ["No data found!!!", "Please refine your search."]

        else:
            context["asset_objects"] = asset_objects
            context["header"] = f"{asset_objects[0].asset_name}"
            context["active_link"] = f"{asset_objects[0].asset_name.asset_name}"
        return context, asset_objects

    def bulk_add_quantity(self, request):
        context = {
            "asset_header": "Bulk Add Quantity",
            "bulk_add_quantity": True,
            "user_fullname": request.user.get_full_name(),
            "form": UploadFileForm(),
        }
        return context

    def post_bulk_assest_quantity_addition(self, request):
        df = pd.read_excel(request.FILES["file"], index_col=None)
        for rows in df.iterrows():
            rows = rows[1]
            # try:
            date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            expiry_date = self.null_check(rows["expiration_date"], date_check=True)

            date_of_purcase = self.null_check(rows["date_of_purchase"], date_check=True)

            asset_name_type = Assets.objects.get(asset_name=rows["asset_name"])

            try:
                facility_object = FacilityDropdown.objects.get(
                    Q(facility_name__icontains=rows["facility"])
                    | Q(facility_code=rows["facility"])
                )
            except Exception as e:
                context = {
                    "error": [
                        f"Bulk Add Asset Was Unsuccessful !!!  ❌ Reason : {e}",
                        "Please Ensure Correct Data is Provided in the Excel File",
                    ]
                }
                return context
            assest_user_obj = None
            current_status = self.null_check(rows["current_status"])
            if self.null_check(rows["asset_user"]):
                current_status = "In Use"
                employee_data = self.get_employee_data(
                    self.null_check(rows["asset_user"])
                )
                if employee_data:
                    assest_user_obj = self.create_or_get_user_without_request(
                        employee_data, rows
                    )
                    ticket_obj = self.bulk_search_ticket_or_create(
                        rows, assest_user_obj, request
                    )

            asset_details = AssetDetails()
            asset_details.asset_name = asset_name_type
            asset_details.brand = self.null_check(rows["brand"])
            asset_details.model_name = self.null_check(rows["model_name"])
            asset_details.model_number = self.null_check(rows["model_number"])
            asset_details.serial_number = self.null_check(rows["serial_number"])
            asset_details.date_of_purchase = date_of_purcase
            asset_details.date_added = date_now
            asset_details.expiration_date = expiry_date
            asset_details.current_status = current_status
            asset_details.description = self.null_check(rows["description"])
            asset_details.facility = facility_object if facility_object else None
            asset_details.added_by = request.user

            if assest_user_obj:
                asset_details.asset_user = assest_user_obj
                asset_details.assign_to_ticket = ticket_obj

            asset_details.save()
        # except Exception as e:
        #     context = {
        #         "error": [
        #             f"Bulk Add Asset Was Unsuccessful !!!  ❌ Reason : {e}",
        #             "Please Ensure Correct Data is Provided in the Excel File",
        #         ]
        #     }
        #     return context

    def null_check(self, pd_row, date_check=False, int_check=False):
        if date_check:
            if pd.notna(pd_row):
                if isinstance(pd_row, datetime):
                    return pd_row
                else:
                    try:
                        return datetime.strptime(pd_row, r"%d-%m-%Y")
                    except:
                        return datetime.strptime(pd_row, r"%d/%m/%Y")
            else:
                return None

        if int_check:
            if pd.notna(pd_row):
                return int(pd_row)
            else:
                return ""

        if pd.notna(pd_row):
            return pd_row
        else:
            return ""

    def get_employee_data(self, employee_name):
        sq = SqlAlchemyConnection()
        search_query = employee_name.split(" ")
        search_string = ""
        for i in range(0, len(search_query)):
            search_string = re.sub(r"[^a-zA-Z\s%]+", "", search_string)
            search_string += f"%{search_query[i]}%"
        employee_data = sq.get_employees_details_with_name(
            search_string.replace(" ", "")
        )
        return employee_data

    def get_employees_details_with_pr_number(self, pr_number, employee_name):
        sq = SqlAlchemyConnection()
        search_query = employee_name.split(" ")
        search_string = ""
        for i in range(0, len(search_query)):
            search_string = re.sub(r"[^a-zA-Z\s%]+", "", search_string)
            search_string += f"%{search_query[i]}%"
        pr_number = re.sub(r"[^0-9]+", "", pr_number)
        employee_data = sq.get_employees_details_with_pr_num_name(
            pr_number, search_string
        )
        return employee_data

    def check_for_NoneType(self, object, index, slice_values):
        try:
            num = object[index][slice_values:]
            if num:
                num = re.sub(r"\D", "", num)
        except:
            num = None
        return num

    def create_or_get_user_without_request(self, employee_data, rows):

        num1 = self.check_for_NoneType(employee_data, 3, -10)
        num2 = self.check_for_NoneType(employee_data, 2, -10)
        full_name = employee_data[0].split()
        first_name = full_name[1:2]  # Get the first two words
        first_name = " ".join(first_name)
        last_name = full_name[2:]  # Get the remaining words
        last_name = " ".join(last_name)  # Join the remaining words with a space
        password = make_password(employee_data[1])
        email = employee_data[4].lower() if employee_data[4] != None else ""

        user, created = User.objects.get_or_create(
            username=employee_data[1],
            defaults={
                "password": password,
                "first_name": first_name,
                "last_name": last_name,
                "email": email,
            },
        )
        
        if created:
            employe_data = Technician(
                user=user,
                department=self.null_check(rows["department"]),
                designation=self.null_check(rows["user_designation"]),
                pr_number=employee_data[1],
                mobile_number=f"{num1}, {num2}",
                extension_number=self.null_check(
                    rows["user_extension"], int_check=True
                ),
            )
            employe_data.save()

        if user:
            if user.first_name == None:
                user.first_name = first_name
            if user.last_name == None:
                user.last_name = last_name
            if user.email == None:
                user.email = employee_data[4].lower()
            user.save()
            try:
                tech = Technician.objects.get(user=user)
            except:
                tech = Technician()
                tech.user =user
            if tech.department == None:
                tech.department = self.null_check(rows["department"])
            if tech.designation == None:
                tech.designation = self.null_check(rows["user_designation"])
            if tech.pr_number == None:
                tech.pr_number = employee_data[1]
            if tech.mobile_number == None:
                tech.mobile_number = f"{num1},{num2}"
            if tech.extension_number == None:
                tech.extension_number = self.null_check(
                    rows["user_extension"], int_check=True
                )
            tech.save()

        
        return user

    def bulk_search_ticket_or_create(self, row, user: User, request):
        ticket = self.null_check(row["assign_to_ticket"])
        if ticket:
            search_ticket_obj = Requests.objects.get(pk=ticket)
            return search_ticket_obj

        else:
            date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            technician, status, req_asin_time = self.bulk_get_technician_and_status(
                date_now, row
            )
            new_serv_req = Requests(
                requester_name=f"{user.first_name} {user.last_name}",
                requester_pr_number=user.technician.pr_number,
                requester_designation=user.technician.designation,
                requester_department=user.technician.department,
                requester_email=user.email,
                requester_extension=user.technician.extension_number,
                requester_phone_number=user.technician.mobile_number,
                request_type="Service Request",
                request_status=status,
                request_mode="Bulk",
                request_priority="Medium",
                request_category="Hardware - Others",
                request_technician=technician,
                subject="✔️ Created with Bulk Asset Creation",
                request_creation_date=date_now,
                request_submitter=request.user,
                last_modified_by=request.user,
                last_modified_date=date_now,
                request_assigned_time=req_asin_time,
            )
            new_serv_req.location = self.null_check(row["location"])
            new_serv_req.save()
            return new_serv_req

    def bulk_get_technician_and_status(self, date_now, row):
        user_techi = self.null_check(row["technician"])
        if user_techi and user_techi.isalnum():
            technician = User.objects.filter(
                Q(username=user_techi)
                | Q(first_name__icontains=user_techi)
                | Q(last_name__icontains=user_techi)
            ).first()

            if (
                technician is not None
                and Group.objects.filter(user=technician, name="Technicians").exists()
            ):
                status = "Assigned"
                req_asin_time = date_now
        else:
            technician = None
            status = "Open"
            req_asin_time = date_now

        return technician, status, req_asin_time

    def send_sms_to_technician(self, technician, requester):
        problem = requester.request_category + " . " + requester.subject
        name = (
            technician.first_name
            + " "
            + technician.last_name
            + " ( "
            + technician.username
            + " )"
        )
        message = default_message.format(
            technician_name=name,
            department=requester.requester_department,
            floor=requester.location,
            requester_name=requester.requester_name,
            phone_number=requester.requester_phone_number,
            extension=requester.requester_extension,
            ticket_id=requester.id,
            problem=problem,
        )
        SendSms(message=message, number=technician.technician.mobile_number)

    def bulk_asset_scrap(self, request):
        context = {
            "asset_header": "Mark Assets As Scrap In Bulk",
            "bulk_asset_scrap": True,
            "user_fullname": request.user.get_full_name(),
        }
        return context

    def post_bulk_asset_scrap(self, request):
        date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        serial_numbers = request.POST.get("bulk_text_area")
        serial_numbers = [serial.strip() for serial in serial_numbers.split("\n")]
        serial_numbers = [serial for serial in serial_numbers if serial]
        status = request.POST.get("current_status_select")
        assets = AssetDetails.objects.filter(serial_number__in=serial_numbers)
        for asset in assets:
            asset.current_status = status
            asset.description += f"Last Modified By {request.user.get_full_name()} ({request.user.username}) On {date_now}. Modification: \n✔️ Changed Current Status Changed To  {status}.\n\n"
            asset.save()

    def reports_tat(request):
        context = {
            "user_fullname": request.user.get_full_name(),
            "header": "TAT Reports",
            "tat_report_page": True,
        }
        ticket_objects = None
        return context, ticket_objects

    def get_tat_report(request):
        # Retrieve POST data
        from_date = datetime.strptime(request.POST.get("from_date"), "%Y-%m-%d").date()
        to_date = datetime.strptime(request.POST.get("to_date"), "%Y-%m-%d").date()
        # Query AssetDetails instances with date_added between from_date and to_date
        queryset = Requests.objects.filter(
            request_creation_date__range=[from_date, to_date + timedelta(days=1)]
        )
        
        # Create an Excel workbook and sheet
        wb = Workbook()

        # Select the active worksheet
        ws = wb.active

        # Define column headers
        headers = [
            "Ticket Number",
            "Requester Name",
            "PR Number",
            "Designation",
            "Department",
            "Email",
            "Extension",
            "Phone Number",
            "Request Type",
            "Request Status",
            "Request Mode",
            "Request Priority",
            "Request Category",
            "Technician",
            "Submitter",
            "Subject",
            "Description",
            "Creation Date",
            "Last Modified By",
            "Last Modified Date",
            "Assigned Time",
            "Resolved Time",
            "Closed Time",
            "Closed User",
            "Location",
        ]

        # Write headers to worksheet
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = Font(bold=True)

        # Write data to worksheet
        for row_num, request in enumerate(queryset, 2):
            row = [
                request.id,
                request.requester_name,
                request.requester_pr_number,
                request.requester_designation,
                request.requester_department,
                request.requester_email,
                request.requester_extension,
                request.requester_phone_number,
                request.request_type,
                request.request_status,
                request.request_mode,
                request.request_priority,
                request.request_category,
                request.request_technician.get_full_name()
                if request.request_technician
                else "",
                request.request_submitter.get_full_name()
                if request.request_submitter
                else "",
                request.subject,
                request.description,
                request.request_creation_date,
                request.last_modified_by.get_full_name()
                if request.last_modified_by
                else "",
                request.last_modified_date,
                request.request_assigned_time,
                request.request_resolved_time,
                request.request_closed_time,
                request.request_closed_user.get_full_name()
                if request.request_closed_user
                else "",
                request.location,
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value

        # Create response object with appropriate content type
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Set content disposition header to trigger a file download
        response["Content-Disposition"] = "attachment; filename=TAT_Report.xlsx"

        # Save workbook data to response
        wb.save(response)

        return response
